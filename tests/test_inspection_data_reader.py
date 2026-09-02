"""
Unit tests for reader.ExcelReader.read_inspection_data() (2026-09-02).

The real Inspection Data workbook has ~150 sheets: one per weekly
window, plus a couple of hand-built summary/tally sheets mixed in at
ARBITRARY positions - not reliably the first sheet, and not
reliably named. These tests build a small synthetic workbook
reproducing that exact shape (a summary sheet first, a second
summary sheet buried in the middle, and data sheets with a couple of
real-world header variations seen in the actual file) and confirm
only the data sheets get combined, harmonized to one fixed schema.
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from constants import INSPECTION_DATA_COLUMNS, INSPECTION_REOFFERED_BEFORE_ACCEPT
from reader import ExcelReader


@pytest.fixture
def reader():
    return ExcelReader()


def _configure(reader, folder: Path):
    reader.settings["paths"]["quality_upload_folder"] = str(folder)
    reader.settings["input_files"]["inspection_data"] = {
        "enabled": True,
        "file_pattern": "*INSPECTION*DATA*.xlsx",
        "header_row": 0,
    }


def _write_workbook(path: Path):
    wb = openpyxl.Workbook()

    # First sheet: a "TYPE OF REWORK" tally, same shape as the real
    # file's Sheet2 - no Project Code/Drawing No/etc at all, so it
    # must never be combined in, regardless of being first.
    summary = wb.active
    summary.title = "Summary"
    summary.append(["TYPE OF REWORK"])
    summary.append(["Bend", "Degree", "Dim"])
    summary.append([3, 3, 3])

    # A normal weekly data sheet - the standard, current column
    # shape used by the overwhelming majority of real sheets.
    week1 = wb.create_sheet("24-08-26 TO 28-08-26")
    week1.append([
        "Project Code", "Drawing No.", "Spool No", "MAT", "Spool Size",
        "Prod. Offer ", "Insp. Remark ", "Final Status ", "Prod. Eng",
    ])
    week1.append([
        "TJ/25-26/182", "DRW-001", "SPOOL-01", "P11", 2,
        "24-08-2026", "Satisfactory", "Accept", "Ankit",
    ])
    week1.append([
        "TJ/25-26/182", "DRW-002", "SPOOL-02", "P11", 1,
        # Re-offer typed into the same cell instead of a new row -
        # the effective date must resolve to the LATEST piece.
        "20-08-2026/26-08-2026", "Bend", "Bend", "Ankit",
    ])
    week1.append([
        "TJ/25-26/182", "DRW-003", "SPOOL-03", "P11", 1,
        "25-08-2026", "Project on hold", "Hold", "Ankit",
    ])
    week1.append([
        "TJ/25-26/182", "DRW-004", "SPOOL-04", "P11", 1,
        # Multi-date offer but Final Status is literally Accept - the
        # real-world pattern the person flagged (2026-09-02): counts
        # as Rework, dated at the EARLIEST piece, not the latest.
        "17-08-2026/20-08-2026/21-08-2026", "tag/punching balance", "Accept", "Ankit",
    ])

    # A second, buried summary sheet - same idea as the real file's
    # stray "Sheet1", sitting between two data sheets rather than at
    # the start, to confirm position never matters, only columns.
    buried_summary = wb.create_sheet("BuriedSummary")
    buried_summary.append(["TYPE OF REWORK"])
    buried_summary.append(["DIM", "BEND"])
    buried_summary.append([5, 19])

    # An older-format sheet: has an extra "Inch Dia" column and is
    # missing "Prod. Eng" entirely - both seen on a handful of real
    # 2023 sheets. Must still be picked up, harmonized down to the
    # standard schema (Inch Dia dropped, Prod Engineer left blank).
    old_format = wb.create_sheet("29-05-23 TO 03-06-23")
    old_format.append([
        "Project Code", "Drawing No.", "Spool No", "MAT", "Inch Dia",
        "Spool Size", "Prod. Offer ", "Insp. Remark ", "Final Status ",
    ])
    old_format.append([
        "TJ/21-22/077", "DRW-OLD", "SPOOL-OLD", "SS", 6,
        3, "29-05-2023", "satisfactory", "Accept",
    ])

    wb.save(path)


def test_summary_sheets_excluded_only_data_sheets_combined(reader, tmp_path):
    """
    Neither the first sheet nor a buried second summary sheet ends up
    in the combined dataframe - only the two real data sheets do (4
    rows from the weekly sheet + 1 row from the old-format sheet).
    """

    _configure(reader, tmp_path)
    _write_workbook(tmp_path / "INSPECTION DATA TEST.xlsx")

    df = reader.read_inspection_data()

    assert df is not None
    assert len(df) == 5
    assert set(df["Spool No"]) == {
        "SPOOL-01", "SPOOL-02", "SPOOL-03", "SPOOL-04", "SPOOL-OLD",
    }


def test_columns_harmonized_to_standard_schema(reader, tmp_path):
    """
    The combined dataframe always has INSPECTION_DATA_COLUMNS plus
    the derived INSPECTION_REOFFERED_BEFORE_ACCEPT flag - the old-
    format sheet's extra "Inch Dia" is dropped, and its missing
    "Prod Engineer" comes back blank rather than raising.
    """

    _configure(reader, tmp_path)
    _write_workbook(tmp_path / "INSPECTION DATA TEST.xlsx")

    df = reader.read_inspection_data()

    assert list(df.columns) == INSPECTION_DATA_COLUMNS + [
        INSPECTION_REOFFERED_BEFORE_ACCEPT
    ]

    old_row = df[df["Spool No"] == "SPOOL-OLD"].iloc[0]
    assert pd.isna(old_row["Prod Engineer"])


def test_multi_date_offer_cell_resolves_to_latest(reader, tmp_path):
    """
    A "/"-separated re-offer cell (the same data-entry pattern the
    Rework Data workbook has) resolves to its LATEST date, same as
    resolve_multi_date_text_cells() already does for that workbook.
    """

    _configure(reader, tmp_path)
    _write_workbook(tmp_path / "INSPECTION DATA TEST.xlsx")

    df = reader.read_inspection_data()

    row = df[df["Spool No"] == "SPOOL-02"].iloc[0]
    assert row["Prod Offer Date"] == pd.Timestamp("2026-08-26")


def test_reoffered_before_accept_flagged_and_dated_earliest(reader, tmp_path):
    """
    A multi-date offer cell whose Final Status is literally "Accept"
    (SPOOL-04) is flagged INSPECTION_REOFFERED_BEFORE_ACCEPT and
    dated at its EARLIEST piece (17-08-2026) - the opposite rule from
    every other multi-date cell (SPOOL-02, latest-wins, tested
    above), since what matters here is when the deficiency was first
    found. A row that's just a plain single-date Accept (SPOOL-01) is
    never flagged.
    """

    _configure(reader, tmp_path)
    _write_workbook(tmp_path / "INSPECTION DATA TEST.xlsx")

    df = reader.read_inspection_data()

    flagged_row = df[df["Spool No"] == "SPOOL-04"].iloc[0]
    assert bool(flagged_row[INSPECTION_REOFFERED_BEFORE_ACCEPT]) is True
    assert flagged_row["Prod Offer Date"] == pd.Timestamp("2026-08-17")

    plain_accept_row = df[df["Spool No"] == "SPOOL-01"].iloc[0]
    assert bool(plain_accept_row[INSPECTION_REOFFERED_BEFORE_ACCEPT]) is False

    bend_row = df[df["Spool No"] == "SPOOL-02"].iloc[0]
    assert bool(bend_row[INSPECTION_REOFFERED_BEFORE_ACCEPT]) is False


def test_missing_workbook_returns_none_not_raise(reader, tmp_path):
    """An empty upload folder is a warning, not a pipeline failure."""

    _configure(reader, tmp_path)

    assert reader.read_inspection_data() is None


def test_disabled_returns_none(reader, tmp_path):
    _configure(reader, tmp_path)
    reader.settings["input_files"]["inspection_data"]["enabled"] = False
    _write_workbook(tmp_path / "INSPECTION DATA TEST.xlsx")

    assert reader.read_inspection_data() is None
