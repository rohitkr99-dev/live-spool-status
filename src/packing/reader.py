"""
src/packing/reader.py
---------------------------------------------------------
Reads every .xlsx workbook in data/upload/packing/ and extracts two
row-level tables per workbook:

  - "spool rows"  - from the sheet whose name contains "spool"
                    (one row per spool)
  - "box rows"    - from the sheet named "Summary" (one row per
                    Box No., with packing/dispatch status)

Header rows aren't at a fixed position (they vary row 1-4 across
files), so both are located by scanning for a recognisable marker
cell rather than assuming a row number.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from packing.logger import logger
from packing.normalize import (
    BOX_COLUMN_ALIASES,
    SPOOL_COLUMN_ALIASES,
    build_header_index,
    normalize_status,
    to_date,
    to_number,
)

# Matches project codes like "TJ/25-26/188", "TJ-25-26-188", "TJ 25-26 188"
PROJECT_CODE_PATTERN = re.compile(r"TJ[\s/-]*\d{2}[\s/-]*\d{2}[\s/-]*\d+", re.IGNORECASE)


def _find_header_row(ws, marker: str, max_scan_rows: int = 6) -> tuple[int, list[Any]] | None:
    """Scan the first max_scan_rows rows for a row containing `marker` exactly (trimmed)."""
    for row_idx in range(1, max_scan_rows + 1):
        values = [cell.value for cell in ws[row_idx]]
        cleaned = [str(v).strip() if v is not None else "" for v in values]
        if marker in cleaned:
            return row_idx, values
    return None


def _find_spool_sheet(wb):
    for ws in wb.worksheets:
        if "spool" in ws.title.lower():
            return ws
    return None


def _find_summary_sheet(wb):
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "summary":
            return ws
    return None


def _guess_bore_type(filename: str) -> str | None:
    lowered = filename.lower()
    if "small bore" in lowered or "small_bore" in lowered:
        return "Small Bore"
    if "large bore" in lowered or "large_bore" in lowered:
        return "Large Bore"
    return None


def _extract_project_title(ws, header_row: int) -> tuple[str | None, str | None]:
    """
    The Summary sheet's project title (e.g. "TJ/25-26/188 / Vogt Power
    ( Bison )") sits somewhere in the rows above the header row.
    Returns (project_code, project_name) parsed out of it, or
    (None, None) if no matching row is found.
    """
    for row_idx in range(1, header_row):
        values = [cell.value for cell in ws[row_idx]]
        for value in values:
            if value is None:
                continue
            text = str(value).strip()
            match = PROJECT_CODE_PATTERN.search(text)
            if match:
                code_raw = match.group(0)
                # Canonical form: TJ/25-26/188
                digits = re.findall(r"\d+", code_raw)
                project_code = f"TJ/{digits[0]}-{digits[1]}/{digits[2]}" if len(digits) >= 3 else code_raw
                name = text[match.end():].strip(" -/")
                # Strip a trailing bore-type suffix so "182" reads as one
                # project name regardless of which split file it came from.
                name = re.sub(r"\s*(Small|Large)\s*Bore\s*$", "", name, flags=re.IGNORECASE).strip()
                return project_code, (name or None)
    return None, None


def read_spool_rows(ws, source_file: str, bore_type: str | None) -> list[dict[str, Any]]:
    found = _find_header_row(ws, "Project Code")
    if not found:
        logger.warning(f"{source_file}: no 'Project Code' header found on sheet '{ws.title}' - skipping spool rows.")
        return []
    header_row, headers = found
    index = build_header_index(headers, SPOOL_COLUMN_ALIASES)

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [cell.value for cell in row]

        def get(field: str):
            i = index.get(field)
            return values[i] if i is not None and i < len(values) else None

        project_code = get("project_code")
        if project_code is None or str(project_code).strip() == "":
            continue

        record = {
            "project_code": str(project_code).strip(),
            "drawing_no": (str(get("drawing_no")).strip() if get("drawing_no") is not None else None),
            "spool_no": (str(get("spool_no")).strip() if get("spool_no") is not None else None),
            "spool_ext_no": (str(get("spool_ext_no")).strip() if get("spool_ext_no") is not None else None),
            "box_no": (str(get("box_no")).strip() if get("box_no") not in (None, "") else None),
            "msn_no": (str(get("msn_no")).strip() if get("msn_no") not in (None, "") else None),
            "paint_system": (str(get("paint_system")).strip() if get("paint_system") not in (None, "") else None),
            "description": (str(get("description")).strip() if get("description") is not None else None),
            "item_category": (str(get("item_category")).strip() if get("item_category") not in (None, "") else None),
            "unit_no": (str(get("unit_no")).strip() if get("unit_no") not in (None, "") else None),
            "total_qty": to_number(get("total_qty")),
            "total_wt": to_number(get("total_wt")),
            "pdi_date": to_date(get("pdi_date")),
            "packing_date": to_date(get("packing_date")),
            "packing_status_raw": (str(get("packing_status")).strip() if get("packing_status") not in (None, "") else None),
            "packing_status": normalize_status(get("packing_status")),
            "dispatched_date": to_date(get("dispatched_date")),
            "inch_dia": to_number(get("inch_dia")),
            "spool_size": to_number(get("spool_size")),
            "spool_type": (str(get("spool_type")).strip() if get("spool_type") not in (None, "") else bore_type),
            "surface_area": to_number(get("surface_area")),
            "handover_date": to_date(get("handover_date")),
            "remark": (str(get("remark")).strip() if get("remark") not in (None, "") else None),
            "source_file": source_file,
        }
        rows.append(record)
    return rows


def read_box_rows(ws, source_file: str, fallback_project_code: str | None) -> list[dict[str, Any]]:
    found = _find_header_row(ws, "Box No.") or _find_header_row(ws, "Box No") or _find_header_row(ws, "Box no")
    if not found:
        logger.info(f"{source_file}: no Summary header row found - skipping box rows.")
        return []
    header_row, headers = found
    index = build_header_index(headers, BOX_COLUMN_ALIASES)
    project_code, project_name = _extract_project_title(ws, header_row)
    project_code = project_code or fallback_project_code

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [cell.value for cell in row]

        def get(field: str):
            i = index.get(field)
            return values[i] if i is not None and i < len(values) else None

        status_raw = get("status")
        box_no = get("box_no")
        if status_raw is None and box_no is None:
            continue
        if status_raw is not None and str(status_raw).strip().upper().startswith("TOTAL"):
            continue

        record = {
            "project_code": project_code,
            "project_name": project_name,
            "box_no": (str(box_no).strip() if box_no not in (None, "") else None),
            "status_raw": (str(status_raw).strip() if status_raw not in (None, "") else None),
            "status": normalize_status(status_raw),
            "qty": to_number(get("qty")),
            "net_wt": to_number(get("net_wt")),
            "item_category": (str(get("item_category")).strip() if get("item_category") not in (None, "") else None),
            "unit_no": (str(get("unit_no")).strip() if get("unit_no") not in (None, "") else None),
            "dispatch_date": to_date(get("dispatch_date")),
            "container_no": (str(get("container_no")).strip() if get("container_no") not in (None, "") else None),
            "seal_no": (str(get("seal_no")).strip() if get("seal_no") not in (None, "") else None),
            "vpi_package_no": (str(get("vpi_package_no")).strip() if get("vpi_package_no") not in (None, "") else None),
            "lbp_sbp": (str(get("lbp_sbp")).strip() if get("lbp_sbp") not in (None, "") else None),
            "remark": (str(get("remark")).strip() if get("remark") not in (None, "") else None),
            "source_file": source_file,
        }
        if record["box_no"] is None:
            continue
        rows.append(record)
    return rows, project_name


def read_workbook(filepath: Path) -> dict[str, Any]:
    """Read one workbook and return {'spools': [...], 'boxes': [...], 'project_name': str|None}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    source_file = filepath.name
    bore_type = _guess_bore_type(source_file)

    spool_sheet = _find_spool_sheet(wb)
    spools: list[dict[str, Any]] = []
    if spool_sheet is not None:
        spools = read_spool_rows(spool_sheet, source_file, bore_type)
    else:
        logger.warning(f"{source_file}: no 'Spool' sheet found.")

    fallback_code = spools[0]["project_code"] if spools else None

    summary_sheet = _find_summary_sheet(wb)
    boxes: list[dict[str, Any]] = []
    project_name = None
    if summary_sheet is not None:
        boxes, project_name = read_box_rows(summary_sheet, source_file, fallback_code)
    else:
        logger.info(f"{source_file}: no 'Summary' sheet found.")

    logger.info(f"{source_file}: read {len(spools)} spool row(s), {len(boxes)} box row(s).")

    return {"spools": spools, "boxes": boxes, "project_name": project_name, "source_file": source_file}


def read_all_workbooks(upload_folder: Path, file_pattern: str = "*.xlsx") -> list[dict[str, Any]]:
    filepaths = sorted(upload_folder.glob(file_pattern))
    if not filepaths:
        logger.warning(f"No workbooks found in {upload_folder} matching {file_pattern}.")
    results = []
    for filepath in filepaths:
        if filepath.name.startswith("~$"):  # Excel lock file
            continue
        try:
            results.append(read_workbook(filepath))
        except Exception as error:  # noqa: BLE001 - one bad workbook shouldn't stop the run
            logger.error(f"Failed to read {filepath.name}: {error}")
    return results
