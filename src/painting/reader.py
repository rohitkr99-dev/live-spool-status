"""
src/painting/reader.py
---------------------------------------------------------
Reads two independent sources and hands both back as plain lists of
dicts - no merging happens here, that's summary.py's job:

  1. read_all_workbooks() - every "Painting Weekly Plan" workbook
     currently in data/upload/painting/ ("Spool List" sheet, one row
     per spool).

  2. read_dpr_rfp_spools() - the Fabrication (DPR) workbook, via the
     SAME ExcelReader the Projects/Production/Packing pipelines all
     read it through (config/settings.json -> input_files.fabrication)
     - best-effort, matching src/packing/pipeline.py ->
     _canonical_project_names()'s pattern: RFP, PDI (= PDI Clearance
     on the DPR - see src/constants.py comment and
     config/production_rules.json -> pdi_clearance_field), Inch Dia,
     Total Wt., Surface Area Out and Quantity are all literal DPR
     Fabrication-sheet columns, no merge-engine cross-referencing
     needed to get them.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from painting.logger import logger
from painting.normalize import (
    SPOOL_COLUMN_ALIASES,
    build_header_index,
    clean_text,
    composite_key,
    to_date,
    to_number,
)


def _find_header_row(ws, marker: str, max_scan_rows: int = 5) -> tuple[int, list[Any]] | None:
    """Scan the first max_scan_rows rows for a row containing `marker` exactly (trimmed)."""
    for row_idx in range(1, max_scan_rows + 1):
        values = [cell.value for cell in ws[row_idx]]
        cleaned = [str(v).strip() if v is not None else "" for v in values]
        if marker in cleaned:
            return row_idx, values
    return None


def _find_spool_list_sheet(wb):
    for ws in wb.worksheets:
        if "spool" in ws.title.lower():
            return ws
    return None


def read_spool_rows(ws, source_file: str) -> list[dict[str, Any]]:
    found = _find_header_row(ws, "Project Code")
    if not found:
        logger.warning(f"{source_file}: no 'Project Code' header found on sheet '{ws.title}' - skipping.")
        return []
    header_row, headers = found
    index = build_header_index(headers, SPOOL_COLUMN_ALIASES)

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):

        def get(field: str):
            i = index.get(field)
            return row[i] if i is not None and i < len(row) else None

        project_code = get("project_code")
        if project_code is None or str(project_code).strip() == "":
            continue

        record = {
            "project_code": str(project_code).strip(),
            "drawing_no": clean_text(get("drawing_no")),
            "spool_no": clean_text(get("spool_no")),
            "paint_system": clean_text(get("paint_system")),
            "item_category": clean_text(get("item_category")),
            "quantity": to_number(get("quantity")),
            "weight": to_number(get("weight")),
            "inch_dia": to_number(get("inch_dia")),
            "surface_area": to_number(get("surface_area")),
            "spool_size": to_number(get("spool_size")),
            "qc_rfp_date": to_date(get("qc_rfp")),
            "painting_plan_week": clean_text(get("painting_plan")),
            "status": clean_text(get("status")),
            "final_remark": clean_text(get("final_remark")),
            "bay_no": clean_text(get("bay_no")),
            "no_of_coats": to_number(get("no_of_coats")),
            "internal_blasting_reqd": clean_text(get("internal_blasting_reqd")),
            "internal_blasting_date": to_date(get("internal_blasting_date")),
            "external_blasting_date": to_date(get("external_blasting_date")),
            "primer_date": to_date(get("primer_date")),
            "mid_coat_1_date": to_date(get("mid_coat_1_date")),
            "mid_coat_2_date": to_date(get("mid_coat_2_date")),
            "top_coat_date": to_date(get("top_coat_date")),
            "pickling_date": to_date(get("pickling_date")),
            "pdi_offer_date": to_date(get("pdi_offer_date")),
            "pdi_status_acceptance_date": to_date(get("pdi_status_acceptance_date")),
            "source_file": source_file,
        }
        record["composite_key"] = composite_key(
            record["project_code"], record["drawing_no"], record["spool_no"]
        )
        rows.append(record)
    return rows


def read_workbook(filepath: Path) -> list[dict[str, Any]]:
    # NOT read_only=True - _find_header_row scans rows via ws[row_idx]
    # and read_spool_rows then does a separate ws.iter_rows() pass;
    # src/packing/reader.py -> read_workbook() uses the same two-pass
    # header-scan-then-iterate technique and also skips read_only for
    # it, so this stays consistent with that rather than risk it.
    wb = openpyxl.load_workbook(filepath, data_only=True)
    source_file = filepath.name

    sheet = _find_spool_list_sheet(wb)
    if sheet is None:
        logger.warning(f"{source_file}: no 'Spool List' sheet found.")
        return []

    rows = read_spool_rows(sheet, source_file)
    logger.info(f"{source_file}: read {len(rows)} spool row(s).")
    return rows


def read_all_workbooks(upload_folder: Path, file_pattern: str = "*.xlsx") -> list[dict[str, Any]]:
    filepaths = sorted(upload_folder.glob(file_pattern))
    if not filepaths:
        logger.warning(f"No workbooks found in {upload_folder} matching {file_pattern}.")
    rows: list[dict[str, Any]] = []
    for filepath in filepaths:
        if filepath.name.startswith("~$"):  # Excel lock file
            continue
        try:
            rows.extend(read_workbook(filepath))
        except Exception as error:  # noqa: BLE001 - one bad workbook shouldn't stop the run
            logger.error(f"Failed to read {filepath.name}: {error}")
    return rows


# ---------------------------------------------------------------
# DPR (Fabrication) cross-reference - which spools have RFP done, and
# their Qty/Weight/Surface Area/Inch Dia/PDI Clearance date, straight
# off the DPR itself.
# ---------------------------------------------------------------

DPR_FIELDS = [
    "Project Code", "Project Name", "Drawing No", "Spool No",
    "Material", "Item Category Code", "Total Qty", "Total Wt.",
    "Inch Dia", "Surface Area Out", "RFP", "PDI",
]


def read_dpr_rfp_spools() -> list[dict[str, Any]]:
    """
    Reads the Fabrication (DPR) workbook via ExcelReader().read_fabrication()
    and returns one dict per spool whose RFP date is filled in - the
    "RFP done" set the Painting pipeline's whole cross-reference is
    built from. Every column pulled is a literal DPR Fabrication-sheet
    field (see module docstring) - no merge-engine involved.

    Returns an empty list (never raises) if the DPR workbook can't be
    read - same best-effort contract as
    src/packing/pipeline.py -> _canonical_project_names(): the
    Painting pipeline still runs, just can't tell which spools are
    RFP-done, and reports that plainly rather than crashing.
    """
    try:
        # Both live at the top level of src/ (sys.path already
        # includes it - see painting_main.py), same as
        # src/packing/pipeline.py -> `from reader import ExcelReader`.
        from reader import ExcelReader  # noqa: E402
        from utils import to_json_safe  # noqa: E402

        fabrication = ExcelReader().read_fabrication()
    except Exception as error:
        logger.warning(
            f"Could not read the Fabrication (DPR) workbook ({error}). "
            "The Painting pipeline can't determine which spools are "
            "RFP-done without it - it will still run, but every spool "
            "will show as 'not found in the DPR'."
        )
        return []

    if fabrication is None or fabrication.empty:
        return []

    missing = [c for c in DPR_FIELDS if c not in fabrication.columns]
    if missing:
        logger.warning(
            f"Fabrication (DPR) workbook is missing expected column(s) "
            f"{missing} - those fields will be blank for every spool "
            "this run."
        )

    rows: list[dict[str, Any]] = []
    for _, row in fabrication.iterrows():
        rfp = to_json_safe(row.get("RFP"))
        if not rfp:
            continue
        project_code = to_json_safe(row.get("Project Code"))
        drawing_no = to_json_safe(row.get("Drawing No"))
        spool_no = to_json_safe(row.get("Spool No"))
        rows.append({
            "composite_key": composite_key(project_code, drawing_no, spool_no),
            "project_code": project_code,
            "project_name": to_json_safe(row.get("Project Name")),
            "drawing_no": drawing_no,
            "spool_no": spool_no,
            "material": to_json_safe(row.get("Material")),
            "item_category": to_json_safe(row.get("Item Category Code")),
            "quantity": to_json_safe(row.get("Total Qty")),
            "weight": to_json_safe(row.get("Total Wt.")),
            "inch_dia": to_json_safe(row.get("Inch Dia")),
            "surface_area": to_json_safe(row.get("Surface Area Out")),
            "rfp_date": rfp,
            "pdi_clearance_date": to_json_safe(row.get("PDI")),
        })

    logger.info(f"Fabrication (DPR): {len(rows)} spool(s) with RFP done.")
    return rows
